import pandas as pd
import requests
import io
from io import BytesIO
import openpyxl
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting import Rule
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime, timedelta


def create_weekly_shopping_list(wb, df, start_date, shopping_day):
    current_date = start_date
    max_weeks = 2  # df['Shopping Period (weeks)'].max()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )
    even_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for week in range(1, max_weeks + 1):
        ws_week = wb.create_sheet(
            title=f"Week {week} ({(current_date + timedelta(weeks=week - 1)).strftime('%Y-%m-%d')})"
        )
        shopping_date = current_date + timedelta(weeks=week - 1)

        # Filter items for the current week, include all items for week 1
        if week == 1:
            week_df = df.copy()
        else:
            week_df = df[
                (df["Shopping Period (weeks)"] == 1)
                | ((week - 1) % df["Shopping Period (weeks)"] == 0)
            ].copy()

        # Remove the URL column for weekly sheets
        week_df_no_url = week_df.drop(columns=["Non Nutrient Data.Image URL"])

        # Add shopping date at the top
        ws_week.append(["Shopping Date:", shopping_date.strftime("%Y-%m-%d")])
        ws_week.append([])

        # Add column headers with the "Status" column
        headers = ["Status"] + list(week_df_no_url.columns)
        ws_week.append(headers)

        # Style the header
        for cell in ws_week[3]:
            cell.font = header_font
            cell.fill = header_fill

        # Add DataFrame to the weekly sheet, shifting columns by one to the right
        for r in dataframe_to_rows(week_df_no_url, index=False, header=False):
            ws_week.append([""] + r)

        # Add images to the weekly sheet
        image_urls_week = week_df["Non Nutrient Data.Image URL"]
        image_row_week = 4  # Start from row 4 (row 1-3 contain headers)

        for image_url in image_urls_week:
            # Download the image
            response = requests.get(image_url)
            image_data = BytesIO(response.content)

            standard_row_height = 15
            ws_week.row_dimensions[image_row_week].height = (
                standard_row_height * 2
            )  # Set the height of the row to twice the standard height

            # Insert the image into the worksheet with offsets
            img = openpyxl.drawing.image.Image(image_data)
            img.width, img.height = (
                img.width * 0.22,
                img.height * 0.22,
            )  # Scale the image
            ws_week.add_image(img, f"A{image_row_week}")

            # Move to the next row
            image_row_week += 1

        # Add data validation for "MISSING" and "GOT IT"
        dv_missing_gotit = DataValidation(
            type="list", formula1='"MISSING,GOT IT"', allow_blank=True
        )
        ws_week.add_data_validation(dv_missing_gotit)

        for row in range(4, len(week_df) + 4):
            cell_number = f"B{row}"
            cell = ws_week[cell_number]
            cell.value = "MISSING"
            dv_missing_gotit.add(cell)

        # Apply alternating row colors
        for row in ws_week.iter_rows(
            min_row=4,
            max_row=len(week_df) + 3,
            min_col=2,
            max_col=len(week_df_no_url.columns) + 1,
        ):
            for cell in row:
                if cell.row % 2 == 0:
                    cell.fill = even_fill

        # Apply dynamic conditional formatting for the "MISSING" and "GOT IT" statuses
        red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )

        red_rule = Rule(type="expression", dxf=DifferentialStyle(fill=red_fill))
        red_rule.formula = ['B4="MISSING"']

        green_rule = Rule(type="expression", dxf=DifferentialStyle(fill=green_fill))
        green_rule.formula = ['B4="GOT IT"']

        ws_week.conditional_formatting.add(f"B4:B{len(week_df) + 3}", red_rule)
        ws_week.conditional_formatting.add(f"B4:B{len(week_df) + 3}", green_rule)

        # Adjust column widths
        for col in ws_week.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = max_length  # + 2
            # ws_week.column_dimensions[column].width = adjusted_width


# Read the CSV file into a DataFrame
def create_shopping_list_sheet(df, shopping_list_columns, output_path):

    df = df[shopping_list_columns]
    df.columns = [col.replace("Shopping List.", "") for col in df.columns]

    # Initialize a new workbook
    wb = Workbook()
    ws_overview = wb.active
    ws_overview.title = "Overview"

    # Create a DataFrame without the URL column for the sheet
    df_no_url = df.drop(columns=["Non Nutrient Data.Image URL"])

    # Add DataFrame to "Overview" sheet, shifting columns by one to the right
    for r_idx, row in enumerate(dataframe_to_rows(df_no_url, index=False, header=True)):
        ws_overview.append([""] + row)

    # Style the header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )

    for cell in ws_overview[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Insert images into column A of the "Overview" sheet
    image_urls = df["Non Nutrient Data.Image URL"]
    image_row = 2  # Start from row 2 (row 1 contains headers)

    for image_url in image_urls:
        # Download the image
        response = requests.get(image_url)
        image_data = BytesIO(response.content)

        standard_row_height = 15
        ws_overview.row_dimensions[image_row].height = (
            standard_row_height * 2
        )  # Set the height of the row to twice the standard height

        # Insert the image into the worksheet with offsets
        img = openpyxl.drawing.image.Image(image_data)
        img.width, img.height = img.width * 0.22, img.height * 0.22  # Scale the image
        ws_overview.add_image(img, f"A{image_row}")

        # Move to the next row
        image_row += 1

    # Add total average weekly price at the bottom
    total_price = df["Average Weekly Price (EUR)"].sum()
    ws_overview.append(["", "", "", "", "", "Total", total_price])

    # Style the total row
    ws_overview[f"F{len(df) + 2}"].font = Font(bold=True)

    # Adjust column widths
    for col in ws_overview.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = max_length  # + 2
        # ws_overview.column_dimensions[column].width = adjusted_width

    # Apply alternating row colors
    even_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for row in ws_overview.iter_rows(
        min_row=2, max_row=len(df) + 1, min_col=2, max_col=len(df_no_url.columns) + 1
    ):
        for cell in row:
            if cell.row % 2 == 0:
                cell.fill = even_fill

    # Save the workbook
    preferred_day = "Monday"
    start_date = datetime.now()

    create_weekly_shopping_list(wb, df, start_date, preferred_day)

    wb_io = io.BytesIO()
    wb.save(wb_io)
    wb_io.seek(0)
    return wb_io
