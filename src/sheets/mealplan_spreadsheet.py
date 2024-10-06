import pandas as pd
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList

# Load the data


def create_mealplan_spreadsheet(optimization_path, spreadsheet_columns, output_path):
    df = pd.read_csv(optimization_path)
    df = df[spreadsheet_columns]

    # Rename columns by dropping the string before the '.'
    df.columns = [col.split(".")[-1] for col in df.columns]
    df.columns = ["Food", "KCAL", "C (g)", "P (g)", "F (g)"]

    # Add a column for Meal Time
    df["Meal Time"] = ""

    # Default assignment function
    def default_assign_meal_time(df):
        meal_times = ["Breakfast", "Lunch", "Dinner"]
        for i in range(len(df)):
            df.at[i, "Meal Time"] = meal_times[i % len(meal_times)]

    # Assign default meal times
    default_assign_meal_time(df)

    # Create a new Excel workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mealplan"

    # Add DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define data validation for Meal Time
    meal_times = ["Breakfast", "Lunch", "Dinner"]
    dv = DataValidation(
        type="list", formula1='"{}"'.format(",".join(meal_times)), allow_blank=True
    )

    # Apply data validation to the Meal Time column
    meal_time_col = ws.max_column
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=meal_time_col)
        dv.add(cell)

    # Add the data validation object to the worksheet
    ws.add_data_validation(dv)

    # Convert the initial data to a table
    table_range = f"A1:{chr(64+meal_time_col)}{ws.max_row}"
    table = Table(displayName="MealplanTable", ref=table_range)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Create a new sheet for separate tables
    ws_separate = wb.create_sheet(title="Meals Breakdown")

    # Function to add a table to the worksheet
    def add_table(ws, df, start_col, title, table_name):
        # Add title
        ws.cell(row=1, column=start_col, value=title).font = Font(bold=True)

        # Add DataFrame headers and rows to the worksheet
        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=3
        ):
            for c_idx, value in enumerate(row, start=start_col):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 2:  # Header row
                    cell.font = Font(bold=True)
                if c_idx == start_col:  # Darker fill for the Food column
                    cell.fill = PatternFill(
                        start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"
                    )
                else:  # Lighter fill for the nutrient columns
                    cell.fill = PatternFill(
                        start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"
                    )

        # Define the range for the table
        table_range = f"{chr(64+start_col)}3:{chr(64+start_col + df.shape[1] - 1)}{df.shape[0] + 3}"

        # Create and style the table
        table = Table(displayName=table_name, ref=table_range)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,  # Use line fashion instead of caro fashion
            showColumnStripes=True,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

        # Add total row
        total_row = df.sum(numeric_only=True).to_frame().T
        total_row["Food"] = "Total"
        total_row_index = df.shape[0] + 3
        for r in dataframe_to_rows(total_row, index=False, header=False):
            for c_idx, value in enumerate(r, start=start_col):
                ws.cell(row=total_row_index + 1, column=c_idx + 1, value=value)

        # Create a bar chart for the totals
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"Total Nutrients for {title}"
        chart.y_axis.title = "Amount"
        chart.x_axis.title = "Nutrients"

        # Focus only on "C (g)", "P (g)", and "F (g)"
        categories = Reference(
            ws, min_col=start_col + 2, min_row=3, max_col=start_col + 4
        )
        chart.set_categories(categories)

        # Add each nutrient series to the chart
        for i, col in enumerate(["C (g)", "P (g)", "F (g)"], start=2):
            data = Reference(
                ws,
                min_col=start_col + i,
                min_row=total_row_index + 1,
                max_row=total_row_index + 1,
            )
            series = Series(data, title=col)
            chart.series.append(series)

        # Add data labels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        # Adjust the position and size of the chart
        chart.width = 8  # Adjusted width
        chart.height = 6  # Adjusted height
        chart_position = f"{chr(64+start_col)}{total_row_index + 3}"
        ws.add_chart(chart, chart_position)

    # Filter data for each meal time
    breakfast_df = df[df["Meal Time"] == "Breakfast"].drop(columns=["Meal Time"])
    lunch_df = df[df["Meal Time"] == "Lunch"].drop(columns=["Meal Time"])
    dinner_df = df[df["Meal Time"] == "Dinner"].drop(columns=["Meal Time"])

    # Calculate the starting columns dynamically
    breakfast_cols = breakfast_df.shape[1]
    lunch_start_col = breakfast_cols + 3
    lunch_cols = lunch_df.shape[1]
    dinner_start_col = lunch_start_col + lunch_cols + 3

    # Add tables to the separate sheet, ensuring data is placed correctly
    ws_separate.cell(row=1, column=1, value="Breakfast").font = Font(bold=True)
    add_table(ws_separate, breakfast_df, 1, "Breakfast", "BreakfastTable")

    ws_separate.cell(row=1, column=lunch_start_col, value="Lunch").font = Font(
        bold=True
    )
    add_table(ws_separate, lunch_df, lunch_start_col, "Lunch", "LunchTable")

    ws_separate.cell(row=1, column=dinner_start_col, value="Dinner").font = Font(
        bold=True
    )
    add_table(ws_separate, dinner_df, dinner_start_col, "Dinner", "DinnerTable")

    # Save the workbook
    directory = os.path.dirname(output_path)

    # Check if the directory exists, if not, create it
    if not os.path.exists(directory):
        os.makedirs(directory)

    # Save the workbook
    wb.save(output_path)
    return wb
